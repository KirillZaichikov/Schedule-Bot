import openpyxl
import os
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message
from aiogram.fsm.context import FSMContext
from Handlers.KB_group import *
from scraper.shedule import WorkWithFile

router = Router()
user_role = "Teacher"
Init_class = WorkWithFile()
Init_class.find_excel_file()
table_file = Init_class.path_to_file
groups = Init_class.get_groups()
print(table_file)

# Функция для загрузки расписания из Excel файла
def load_schedule_from_excel(file_path):
    if not os.path.isfile(file_path):
        print("Ошибка: файл не найден. Убедитесь, что файл существует.")
        return {}  # Возвращаем пустой словарь, если файл не найден
    schedule = {}
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Получаем активный лист
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            if row[0] is None or row[1] is None:  # Проверка на наличие данных
                continue
            day = row[0]
            time = row[1]  # Часы
            subject = row[2]  # Предмет
            room = row[3]  # Аудитория
            teacher = row[4] if row[4] is not None else "Неизвестный преподаватель"  # Заменяем None на текст
            if day not in schedule:
                schedule[day] = []
            schedule[day].append((time, subject, room, teacher))  # Сохраняем данные по дням
    except Exception as e:
        print(f"Ошибка при загрузке расписания: {e}")
    return schedule


# Загружаем расписание из файла
schedule = load_schedule_from_excel(table_file)

@router.message(Command("Couples"))
async def teacher_button(message: Message, state: FSMContext):
    user_data = await state.get_data()  # Получаем данные о пользователе из состояния
    await message.answer("Теперь вы можете посмотреть расписание любой группы!!", reply_markup=kb())


@router.message(F.text.in_(groups))  # groups - это список
async def button(message: Message, state: FSMContext):
    global table_file  # Указываем, что используем глобальную переменную
    group_name = message.text  # Получаем название группы

    # Загружаем рабочую книгу
    try:
        workbook = openpyxl.load_workbook(table_file)  # Используем правильный путь к файлу
        if group_name not in workbook.sheetnames:
            await message.answer("Группа не найдена.")
            return
        sheet = workbook[group_name]
    except Exception as e:
        await message.answer(f"Ошибка при открытии файла: {str(e)}")
        return

    # Сбор данных для расписания
    schedule_message = "Вот ваше расписание:\n\n"
    unique_entries = set()  # Множество для хранения уникальных записей
    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Получаем все строки в виде списка

    for i, row in enumerate(rows):
        if row[0] is None:  # Проверка на наличие данных в первой ячейке
            continue
        day, time, subject, room = row[:4]  # Убираем преподавателя из распаковки

        # Создаем уникальный ключ для записи
        entry_key = (day, time, subject, room)
        if entry_key not in unique_entries:  # Проверяем, есть ли уже такая запись
            unique_entries.add(entry_key)  # Добавляем уникальную запись

            # Формируем строку с расписанием
            schedule_message += f"{day} - {time}: {subject}, \"{room} Аудитория\"\n\n"  # Добавляем пробелы

            # Проверяем, есть ли следующая строка и добавляем ФИО преподавателя
            if i + 1 < len(rows):  # Если следующая строка существует
                next_row = rows[i + 1]
                if next_row[0] is None:  # Проверяем, пустая ли первая ячейка
                    teacher_name = next_row[1]  # Предполагаем, что ФИО преподавателя находится во второй ячейке
                    if teacher_name:  # Если ФИО не пустое
                        schedule_message += f"\t{teacher_name}\n\n"  # Добавляем пробелы

    # Проверяем, есть ли данные для отображения
    if not schedule_message.strip():
        await message.answer("Нет данных для отображения.")
        return

    # Удаляем сообщение пользователя с выбором группы
    try:
        await message.delete()  # Удаляем сообщение пользователя
    except Exception as e:
        print(f"Ошибка при удалении сообщения пользователя: {e}")

    # Получаем идентификатор старого сообщения бота из состояния
    user_data = await state.get_data()
    last_message_id = user_data.get("last_message_id")

    # Удаляем старое сообщение бота, если оно существует
    if last_message_id:
        try:
            await message.bot.delete_message(chat_id=message.chat.id, message_id=last_message_id)
        except Exception as e:
            print(f"Ошибка при удалении старого сообщения бота: {e}")

    # Отправляем новое сообщение с расписанием
    new_message = await message.answer(schedule_message)

    # Сохраняем идентификатор нового сообщения в состоянии
    await state.update_data(last_message_id=new_message.message_id)






# 1 Создать функционал работы с преподавателем.
# 1.1 Создать список групп в виде кнопок, вывод названия группы. ✔
# 1.2  Обеспечить работоспособность вывода расписания групп. +-✔
# 1.3 Обеспечить форматирование внешнего вида таблицы с расписанием. ✔
# 1.4 Создать клавишу "Мои Пары" первоначальный вывод "Ваши группы".
# 1.5 Создать функцию парсинга пар по имени преподавателя.
# 1.6 Связать вывод "Ваши группы" + задача 1.5.
# Разрешить изменения вывода текста.