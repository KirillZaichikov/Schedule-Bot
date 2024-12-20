import sqlite3
from datetime import datetime

from openpyxl import load_workbook
import pathlib
import sys


class GroupOfFiles:
    list_of_files = []
    def __init__(self):
        path = pathlib.Path(r"C:\Users\user\Documents\ВУЗ\Schedule-Bot\scraper")
        print(path.glob("*.xlsx"))
        for i in path.glob("*.xlsx"):
            self.list_of_files.append(i)


class WorkWithFile:
    def __init__(self):
        self.path_to_file = ''

    def find_excel_file(self):
        path = (pathlib.Path(__file__)).parents[0]
        for i in path.glob("*.xlsx"):
            self.path_to_file = i
            break

    def get_groups(self):
        if self.path_to_file:
            wb = load_workbook(self.path_to_file)
            return wb.sheetnames
        else:
            print("Excel-файл не найден!")


class Date:
    def TodayDate(self):
        today = datetime.now()

        # Форматируем дату
        formatted_date = today.strftime("%d.%m")

        # Заменяем сокращение дня недели на нужное
        day_of_week = today.strftime("%a")
        if day_of_week == "Fri":
            day_of_week = "ПТ"
        elif day_of_week == "Mon":
            day_of_week = "ПН"
        elif day_of_week == "Tue":
            day_of_week = "ВТ"
        elif day_of_week == "Wed":
            day_of_week = "СР"
        elif day_of_week == "Thu":
            day_of_week = "ЧТ"
        elif day_of_week == "Sat":
            day_of_week = "СБ"
        elif day_of_week == "Sun":
            day_of_week = "ВС"

        return f"{formatted_date}\n{day_of_week}"


def ReadGroupShedule(path_to_file, group): # чтение без корректировок
        workbook = load_workbook(path_to_file)
        sheet = workbook[group]
        return sheet


def WorkWithText(group, path):
    x = 1
    text = ""
    day = []
    try:
        shedule = ReadGroupShedule(path_to_file=path, group=group)
    except Exception as e:
        print(f"Ошибка при чтении расписания: {e}")
    for rows in shedule.iter_rows(min_row=16):
        for row in rows:
            if row.value is not None:  # Проверяем, что значение не None
                if x < 2:
                    text += f"\n{row.value}"
                    x += 1
                else:
                    if row.coordinate[0] == "A":  # Проверяем, содержится ли день недели в строке
                        day.append(text)
                        text = ""
                        pass
                    text += f"\n{row.value}"
    return day

def table(shedule, group):
    con = sqlite3.connect("Test_db.db")
    cur = con.cursor()
    group = group.replace(" ", "_")
    group = group.replace("-", "_")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS {group}(
    Date TEXT,
    Day_in_week TEXT,
    shed TEXT)""")
    for Shedule in shedule:
        data = Shedule[1:7]
        Day_in_Week = Shedule[7:10]
        cur.execute(f"""INSERT INTO {group}(Date, Day_in_week, shed) VALUES(?, ?, ?)""",
                    (data, Day_in_Week, Shedule))
    con.commit()
    con.close()